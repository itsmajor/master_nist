#!/usr/bin/env python3
import os
import shutil
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# 1. Pfade definieren
csv_input = 'output/csv/nist_all.csv'
xlsx_dir  = 'output/xlsx/nist'
combined_file = os.path.join(xlsx_dir, 'nist_all.xlsx')

# 2. Output-Verzeichnis bereinigen
if os.path.exists(xlsx_dir):
    shutil.rmtree(xlsx_dir)
os.makedirs(xlsx_dir, exist_ok=True)

# 3. CSV einlesen
df = pd.read_csv(csv_input)
# 4. Runde fix auf 5 setzen (jeder Algorithmus nutzt R=5)
round_value = 5

# 5. Flaches DataFrame ohne wiederholte Model-Spalten
df_flat = pd.DataFrame({
    'Model':      df['Model'],
    'Key':        df['keygen_time'],
    'Enc':        df['enc_time'],
    'Dec':        df['dec_time'],
    'Key2':       df['keygen_heap'],
    'Enc2':       df['enc_heap'],
    'Dec2':       df['dec_heap'],
    'Key3':       df['keygen_stack'],
    'Enc3':       df['enc_stack'],
    'Dec3':       df['dec_stack'],
})
# Spaltenbezeichnungen anpassen für Excel (zweite Header-Zeile)
col_labels = ['Model', 'Key', 'Enc', 'Dec', 'Key', 'Enc', 'Dec', 'Key', 'Enc', 'Dec']
df_flat.columns = col_labels

# Gruppierungen für Header (1-basierter Index)
group_ranges = [('Time', 2, 4), ('Heap', 5, 7), ('Stack', 8, 10)]
# Untertitel für Header-Zeile 2
subtitles = ['Model'] + ['Key','Enc','Dec'] * 3

# 6. Kombiniertes XLSX schreiben
with pd.ExcelWriter(combined_file, engine='openpyxl') as writer:
    df_flat.to_excel(writer, index=False, header=False, sheet_name='NIST All', startrow=2)
    wb = writer.book
    ws = writer.sheets['NIST All']

    # 7. Obere Zelle A1 = Runden
    ws.cell(row=1, column=1, value=f"R={round_value}")

    # 8. Erste Header-Zeile (Gruppentitel)
    for title, start, end in group_ranges:
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        c = ws.cell(row=1, column=start)
        c.value = title
        c.font = Font(name='Arial', size=8, bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center')

    # 9. Zweite Header-Zeile (Untertitel)
    for idx, sub in enumerate(subtitles, start=1):
        c = ws.cell(row=2, column=idx, value=sub)
        c.font = Font(name='Arial', size=8, bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center')

    # 10. Zellformatierung: Schrift Arial 8 & Rahmen
    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=10):
        for cell in row:
            cell.font = Font(name='Arial', size=8)
            cell.border = border

    # 11. Spaltenbreiten an Inhalt anpassen
    for col in range(1, 11):
        letter = get_column_letter(col)
        max_len = max((len(str(c.value)) for c in ws[letter] if c.value is not None), default=0)
        ws.column_dimensions[letter].width = max_len + 2

# 12. Einzelne XLSX pro Algorithmus erstellen
for algo, grp in df.groupby('Algorithmus'):
    safe = algo.replace('/', '_').replace(' ', '_')
    file_path = os.path.join(xlsx_dir, f'{safe}.xlsx')
    # neues Workbook
    wb_algo = Workbook()
    ws_algo = wb_algo.active
    ws_algo.title = 'NIST All'

    # Header-Zeilen identisch zur kombinierten Datei
    ws_algo.cell(row=1, column=1, value=f"R={round_value}")
    for title, start, end in group_ranges:
        ws_algo.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        c = ws_algo.cell(row=1, column=start)
        c.value = title
        c.font = Font(name='Arial', size=8, bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center')
    for idx, sub in enumerate(subtitles, start=1):
        c = ws_algo.cell(row=2, column=idx, value=sub)
        c.font = Font(name='Arial', size=8, bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center')

    # Daten ab Zeile 3
    df_algo_flat = grp[['Model'] + cols[1:]].copy()
    df_algo_flat.columns = col_labels
    for r, row in enumerate(df_algo_flat.itertuples(index=False), start=3):
        for c_idx, val in enumerate(row, start=1):
            cell = ws_algo.cell(row=r, column=c_idx, value=val)
            cell.font = Font(name='Arial', size=8)
            cell.border = border

    # Spaltenbreiten anpassen für Algo-Datei
    for col in range(1, 11):
        letter = get_column_letter(col)
        max_len = max((len(str(c.value)) for c in ws_algo[letter] if c.value is not None), default=0)
        ws_algo.column_dimensions[letter].width = max_len + 2

    wb_algo.save(file_path)
    print(f"Einzeldatei für {algo}: {file_path}")

print(f"Kombinierte Datei erstellt: {combined_file}")
